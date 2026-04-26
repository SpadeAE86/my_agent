"""
Extract distinct label sets from a metadata Excel file.

Goal:
- Collect unique values (as "label sets") for selected columns.
- Write results to a JSON file for later prompt/schema constraint design.

Notes:
- This script is intentionally "PyCharm friendly": no argparse required.
- It uses openpyxl (avoid pandas/pyarrow dependency issues).

Run:
  python -m src.test.extract_xlsx_tag_sets
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple


# === Configure here ===
XLSX_PATH = Path(r"C:\Users\admin\Downloads\LS6视频\数字人素材（LS9、全新L6）.xlsx")

# Columns the user explicitly cares about (as "keyword fields"):
TARGET_COLUMNS: List[str] = [
    "车型",
    "尺寸",
    "画面类型",
    "画面场景",
    "车色",
    "产品状态场景",
    "产品功能场景",
    "产品细节",
    "是否含达人肖像",
]

# Output JSON path (inside repo for easy diff/iteration)
OUTPUT_JSON = Path(__file__).resolve().parent / "sample" / "xlsx_tag_sets.json"

# If a cell contains multiple labels, split by these separators.
_SPLIT_RE = re.compile(r"[、,，;；/|｜\n\r\t]+")


def _to_text(v: Any) -> str:
    if v is None:
        return ""
    # openpyxl may return numbers / booleans / formulas; normalize.
    if isinstance(v, bool):
        return "是" if v else "否"
    return str(v).strip()


def _split_multi_value_cell(text: str) -> List[str]:
    t = text.strip()
    if not t:
        return []
    parts = [p.strip() for p in _SPLIT_RE.split(t)]
    # Drop empty / placeholder values commonly seen in sheets.
    bad = {"", "-", "—", "无", "暂无", "N/A", "NA", "null", "None"}
    return [p for p in parts if p and p not in bad]


def _load_header_and_rows(xlsx_path: Path) -> Tuple[List[str], Iterable[List[Any]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:  # pragma: no cover
        raise RuntimeError(
            "Missing dependency: openpyxl. Install it with: pip install openpyxl"
        ) from e

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        raise RuntimeError(f"Empty sheet: {xlsx_path}")

    header = [_to_text(x) for x in header_row]
    return header, rows


def _build_col_index(header: List[str]) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for i, name in enumerate(header):
        if name and name not in mapping:
            mapping[name] = i
    return mapping


def main() -> None:
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"XLSX not found: {XLSX_PATH}")

    header, rows = _load_header_and_rows(XLSX_PATH)
    col_index = _build_col_index(header)

    missing = [c for c in TARGET_COLUMNS if c not in col_index]
    if missing:
        raise RuntimeError(
            "Some target columns are missing in the first sheet header.\n"
            f"Missing: {missing}\n"
            f"Available: {header}"
        )

    sets: Dict[str, Set[str]] = {c: set() for c in TARGET_COLUMNS}
    non_empty_counts: Dict[str, int] = {c: 0 for c in TARGET_COLUMNS}
    row_count = 0

    for row in rows:
        row_count += 1
        for col in TARGET_COLUMNS:
            idx = col_index[col]
            v = row[idx] if idx < len(row) else None
            text = _to_text(v)
            if not text:
                continue
            non_empty_counts[col] += 1
            for item in _split_multi_value_cell(text):
                sets[col].add(item)

    # Sort for stable JSON diffs.
    result: Dict[str, Any] = {
        "source_xlsx": str(XLSX_PATH),
        "sheet": "0",
        "total_data_rows": row_count,
        "columns": {
            col: {
                "unique_count": len(sets[col]),
                "non_empty_cell_count": non_empty_counts[col],
                "unique": sorted(sets[col]),
            }
            for col in TARGET_COLUMNS
        },
    }

    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote: {OUTPUT_JSON}")


if __name__ == "__main__":
    main()

