import os
import re
import requests
from openpyxl import load_workbook
from urllib.parse import urlparse, unquote

xlsx_path = r"C:\Users\25065\Documents\xwechat_files\wxid_mjgwxtms1uy722_dde3\temp\RWTemp\2026-04\21c0950a70e2959de8e182c102092bc7\数字人素材（LS9、全新L6）.xlsx"
output_dir = "./downloads"
sheet_name = None

os.makedirs(output_dir, exist_ok=True)

# 提取 HYPERLINK 公式里的 URL
hyperlink_pattern = re.compile(r'HYPERLINK\("([^"]+)"', re.IGNORECASE)

def extract_url(cell):
    """
    兼容三种情况：
    1. Excel hyperlink对象
    2. =HYPERLINK("url","name")公式
    3. 纯文本URL
    """
    # 1. Excel真正的超链接对象
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target

    val = cell.value

    if not val:
        return None

    if isinstance(val, str):
        # 2. 公式形式 =HYPERLINK("url","name")
        match = hyperlink_pattern.search(val)
        if match:
            return match.group(1)

        # 3. 纯URL
        if val.startswith("http"):
            return val

    return None


def get_filename(url):
    parsed = urlparse(url)
    name = os.path.basename(parsed.path)
    name = unquote(name)

    if not name:
        name = "file_" + str(abs(hash(url)))

    return name


def download(url, path):
    try:
        r = requests.get(url, stream=True, timeout=30)
        r.raise_for_status()

        with open(path, "wb") as f:
            for chunk in r.iter_content(8192):
                if chunk:
                    f.write(chunk)

        print(f"✔ {path}")

    except Exception as e:
        print(f"✘ 下载失败: {url} -> {e}")


def main():
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]

        url = extract_url(cell)

        if not url:
            continue

        filename = get_filename(url)
        save_path = os.path.join(output_dir, filename)

        download(url, save_path)


if __name__ == "__main__":
    main()